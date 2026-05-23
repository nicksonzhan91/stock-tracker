"""
export_value_excel.py
從永豐金抓取通過 7 年篩選的 45 支股票資料，匯出 Excel（兩個工作表）

工作表 1：合理估值
  ・近 5 年最高 PE（zca 頁面預計算值，年度最高本益比的最大值）
  ・近 5 年最低 PE（zca 頁面預計算值，年度最低本益比的最小值）
  ・最近年度每股淨值(F)（zcra 頁面）
  ・合理價值 = 每股淨值 × (高PE + 低PE) / 2 × 近5年平均ROE(小數)

工作表 2：ROE & 歸母淨利（近 8 年）

用法：
  python scripts/export_value_excel.py
"""

import json, os, sys, time, re, requests, urllib3
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

urllib3.disable_warnings()
sys.stdout.reconfigure(encoding="utf-8")

# ── 路徑 ──
VALUE_LIST  = os.path.join(os.path.dirname(__file__), "..", "data", "value_list.json")
OUTPUT_XLS  = os.path.join(os.path.dirname(__file__), "..", "data", "value_stocks.xlsx")
OUTPUT_ZONE = os.path.join(os.path.dirname(__file__), "..", "data", "value_zone.json")

SINOTRADE = "https://stockchannelnew.sinotrade.com.tw"
_HEADERS  = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
             "Accept-Language": "zh-TW,zh;q=0.9"}
SLEEP     = 0.3
ALL_YEARS = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]   # 近 8 年
PE_YEARS  = [2021, 2022, 2023, 2024, 2025]                      # 近 5 年（PE & ROE 平均）


# ════════════════════════════════════════════════════════
#  永豐金資料抓取
# ════════════════════════════════════════════════════════

def _parse_row(html_bytes_or_str, label_keyword: str) -> dict[int, float | None]:
    """解析 div.table-row 頁面，找到包含 label_keyword 的列，回傳 {year: value}"""
    if isinstance(html_bytes_or_str, bytes):
        soup = BeautifulSoup(html_bytes_or_str, "html.parser")
    else:
        soup = BeautifulSoup(html_bytes_or_str, "html.parser")

    rows = soup.find_all("div", class_="table-row")
    year_row = target_row = None

    for row in rows:
        cells = [s.get_text(strip=True)
                 for s in row.find_all("span", class_=re.compile(r"table-cell"))]
        if not cells:
            continue
        if any(c.isdigit() and len(c) == 4 for c in cells):
            year_row = cells
        if cells and label_keyword in cells[0]:
            target_row = cells

    if not year_row or not target_row:
        return {}

    result = {}
    for i, y in enumerate(year_row):
        if y.isdigit() and len(y) == 4 and i < len(target_row):
            try:
                result[int(y)] = float(target_row[i].replace(",", ""))
            except ValueError:
                result[int(y)] = None
    return result


def fetch_zcra(stock_id: str) -> dict[str, dict[int, float | None]]:
    """從 zcra 一次抓 ROE(A)、每股淨值(F)"""
    url = f"{SINOTRADE}/z/zc/zcr/zcra/zcra_{stock_id}.djhtm"
    try:
        r = requests.get(url, headers=_HEADERS, timeout=15, verify=False)
        content = r.content
        return {
            "roe": _parse_row(content, "ROE(A)"),
            "bv":  _parse_row(content, "每股淨值(F)"),
        }
    except Exception:
        return {"roe": {}, "bv": {}}




def fetch_income(stock_id: str) -> dict[int, float | None]:
    """回傳 {year: 億元}（原始單位百萬元，除以 100）"""
    url = f"{SINOTRADE}/z/zc/zcq/zcqa/zcqa_{stock_id}.djhtm"
    try:
        r = requests.get(url, headers=_HEADERS, timeout=15, verify=False)
        raw = _parse_row(r.content, "歸屬母公司淨利")
        return {y: round(v / 100, 2) if v is not None else None
                for y, v in raw.items()}
    except Exception:
        return {}


def fetch_zca_data(stock_id: str) -> dict:
    """從 zca 一次抓最高/最低本益比 + 前日收盤價（全部來自永豐金，Big5 HTML）

    PE 表格（7 列固定結構）：
      Row 0: 年度（民國年，如 115=2025, 114=2024 …）
      Row 1: 最高成交值
      Row 2: 最低成交值
      Row 3: 最高本益比  ← 使用此列
      Row 4: 最低本益比  ← 使用此列
      Row 5: 盈餘配股
      Row 6: 現金股利

    收盤價（前日收盤）：找第一個有 8 欄且欄 1,3,5,7 均為正數的列
      col 1 = 現價, col 3 = 日高, col 5 = 日低, col 7 = 前日收盤  ← 使用此欄

    回傳 {'high_pe': {西元年: val}, 'low_pe': {西元年: val}, 'price': float|None}
    """
    url = f"{SINOTRADE}/z/zc/zca/zca_{stock_id}.djhtm"
    result = {"high_pe": {}, "low_pe": {}, "price": None}
    try:
        r = requests.get(url, headers=_HEADERS, timeout=15, verify=False)
        html = r.content.decode("big5", errors="replace")
        soup = BeautifulSoup(html, "html.parser")

        # ── 1. 抓收盤價：找欄 1,3,5,7 均為正數的列 ──
        for table in soup.find_all("table"):
            for row in table.find_all("tr"):
                cells = [td.get_text(strip=True) for td in row.find_all(["th", "td"])]
                if len(cells) < 8:
                    continue
                try:
                    vals = [float(cells[i].replace(",", "")) for i in [1, 3, 5, 7]]
                    if all(1 < v < 100000 for v in vals):
                        result["price"] = vals[3]   # col 7 = 前日收盤
                        break
                except (ValueError, IndexError):
                    pass
            if result["price"] is not None:
                break

        # ── 2. 抓 PE：固定 7 列表格 + 民國年 105–120 ──
        for table in soup.find_all("table"):
            rows = table.find_all("tr")
            if len(rows) != 7:
                continue

            header = [td.get_text(strip=True) for td in rows[0].find_all(["th", "td"])]
            year_map: dict[int, int] = {}
            for i, c in enumerate(header):
                if c.isdigit() and 105 <= int(c) <= 120:
                    year_map[i] = int(c) + 1911

            if len(year_map) < 4:
                continue

            high_pe: dict[int, float | None] = {}
            low_pe:  dict[int, float | None] = {}
            high_row = [td.get_text(strip=True) for td in rows[3].find_all(["th", "td"])]
            low_row  = [td.get_text(strip=True) for td in rows[4].find_all(["th", "td"])]

            for idx, yr in year_map.items():
                for res, row_data in [(high_pe, high_row), (low_pe, low_row)]:
                    if idx < len(row_data):
                        val_str = row_data[idx].replace(",", "")
                        try:
                            v = float(val_str)
                            res[yr] = v if v > 0 else None
                        except ValueError:
                            res[yr] = None

            if high_pe or low_pe:
                result["high_pe"] = high_pe
                result["low_pe"]  = low_pe
                break

    except Exception:
        pass
    return result


def calc_pe_metrics(high_pe: dict, low_pe: dict, years: list) -> dict:
    """從 zca 預計算的高/低 PE 字典中，取近 N 年的最高/最低值
    回傳 {'max_pe': float, 'min_pe': float, 'n_years': int}
    """
    hi_vals = {y: high_pe[y] for y in years if high_pe.get(y) is not None}
    lo_vals = {y: low_pe[y]  for y in years if low_pe.get(y)  is not None}

    max_pe = round(max(hi_vals.values()), 2) if hi_vals else None
    min_pe = round(min(lo_vals.values()), 2) if lo_vals else None

    return {
        "max_pe":  max_pe,
        "min_pe":  min_pe,
        "n_years": len(hi_vals),
    }


def calc_value(bv: float | None, max_pe: float | None,
               min_pe: float | None, avg_roe_pct: float | None) -> float | None:
    """合理價值 = 每股淨值 × (近5年最高PE + 近5年最低PE) / 2 × 近5年平均ROE(小數)"""
    if any(v is None for v in [bv, max_pe, min_pe, avg_roe_pct]):
        return None
    return round(bv * (max_pe + min_pe) / 2 * (avg_roe_pct / 100), 2)


# ════════════════════════════════════════════════════════
#  Excel 樣式
# ════════════════════════════════════════════════════════

def _clr(hex_str):
    return PatternFill("solid", fgColor=hex_str)

def _border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL   = _clr("1565C0")   # 深藍  — 主標題
YR_FILL    = _clr("1976D2")   # 藍    — 欄位標題
ROE_FILL   = _clr("E3F2FD")   # 淡藍  — ROE 數值背景
INC_FILL   = _clr("E8F5E9")   # 淡綠  — 淨利數值背景
WARN_FILL  = _clr("FFF9C4")   # 淡黃  — ROE < 15% 警示
VAL_FILL   = _clr("FFF3E0")   # 淡橘  — 估值工作表數值背景
GOOD_FILL  = _clr("C8E6C9")   # 深綠  — 現價低於合理價值
WHITE_FILL = _clr("FFFFFF")

HDR_FONT   = Font(bold=True, color="FFFFFF", name="微軟正黑體", size=11)
YR_FONT    = Font(bold=True, color="FFFFFF", name="微軟正黑體", size=10)
LABEL_FONT = Font(bold=True, name="微軟正黑體", size=10)
DATA_FONT  = Font(name="Calibri", size=10)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center")
RIGHT      = Alignment(horizontal="right",  vertical="center")


# ════════════════════════════════════════════════════════
#  工作表 1：合理估值
# ════════════════════════════════════════════════════════

def build_valuation_sheet(ws, all_data: list):
    ws.title = "合理估值"
    ws.freeze_panes = "C3"

    # 欄寬
    col_widths = [8, 12, 9, 9, 9, 9, 9, 9, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    n_cols = 9

    # 第 1 列：大標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1,
                   value="價值型股票 — 合理估值（近5年高/低PE × 每股淨值 × ROE）")
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # 第 2 列：欄位標題
    headers = ["代號", "名稱",
               f"近5年\n最高PE", f"近5年\n最低PE",
               f"近5年\n平均ROE%",
               "每股淨值(F)\n(最新年度)",
               "合理價值\n(元)",
               "平均PE\n參考",
               "計算年數\n(PE)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font      = YR_FONT
        cell.fill      = YR_FILL
        cell.alignment = CENTER
        cell.border    = _border()
    ws.row_dimensions[2].height = 30

    # 資料列
    for row_idx, d in enumerate(all_data, 3):
        sid       = d["stock_id"]
        name      = d["name"]
        max_pe    = d["max_pe"]
        min_pe    = d["min_pe"]
        avg_roe   = d["avg_roe_pct"]
        bv_latest = d["bv_latest"]
        value     = d["value"]
        avg_pe    = round((max_pe + min_pe) / 2, 2) if (max_pe and min_pe) else None
        n_pe_yrs  = d["n_pe_years"]

        row_data = [sid, name, max_pe, min_pe, avg_roe, bv_latest, value, avg_pe, n_pe_yrs]
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.border    = _border()
            cell.alignment = CENTER if c <= 2 else RIGHT
            cell.font      = LABEL_FONT if c <= 2 else DATA_FONT
            if c <= 2:
                cell.fill = VAL_FILL
            elif c == 7:   # 合理價值 — 數字格式
                cell.number_format = "#,##0.00"
                cell.fill          = GOOD_FILL
            elif c in (3, 4, 8):  # PE
                cell.number_format = "0.00"
                cell.fill          = VAL_FILL
            elif c == 5:   # ROE%
                cell.number_format = "0.00"
                cell.fill          = VAL_FILL
            elif c == 6:   # 每股淨值
                cell.number_format = "0.00"
                cell.fill          = VAL_FILL
            else:
                cell.fill = WHITE_FILL

        ws.row_dimensions[row_idx].height = 18

    # 底部說明
    note_row = len(all_data) + 4
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=n_cols)
    ws.cell(row=note_row, column=1,
            value=("※ 合理價值公式：每股淨值(F) × (近5年最高PE + 近5年最低PE) / 2 × 近5年平均ROE(小數)"
                   "    ※ 最高/最低 PE 來源：永豐金 zca 頁面預計算值（年度最高/最低本益比）"
                   "    ※ 資料來源：永豐金股市資訊網")
            ).font = Font(italic=True, color="757575", name="微軟正黑體", size=9)


# ════════════════════════════════════════════════════════
#  工作表 2：ROE & 歸母淨利（近 8 年）
# ════════════════════════════════════════════════════════

def build_detail_sheet(ws, all_data: list):
    ws.title = "ROE & 歸母淨利（近8年）"
    ws.freeze_panes = "D3"

    # 欄寬
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    for col_idx in range(4, 4 + len(ALL_YEARS)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9

    n_cols = 3 + len(ALL_YEARS)

    # 第 1 列：大標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1,
                   value="價值型股票篩選結果（近7年 ROE≥15% 且 歸母淨利≥5億）— ROE(A) & 歸母淨利 近8年")
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # 第 2 列：欄位標題
    headers = ["代號", "名稱", "指標"] + [str(y) for y in ALL_YEARS]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font      = YR_FONT
        cell.fill      = YR_FILL
        cell.alignment = CENTER
        cell.border    = _border()
    ws.row_dimensions[2].height = 20

    # 資料列（每股票 2 列：ROE + 淨利）
    row = 3
    for d in all_data:
        sid  = d["stock_id"]
        name = d["name"]
        roe  = d["roe"]
        inc  = d["income"]

        # ROE 列
        for c, val in enumerate([sid, name, "ROE(A) %"], 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.alignment = CENTER
            cell.font      = LABEL_FONT
            cell.border    = _border()
            cell.fill      = ROE_FILL

        # 淨利列
        for c, val in enumerate([sid, name, "歸母淨利 億"], 1):
            cell = ws.cell(row=row+1, column=c, value=val)
            cell.alignment = CENTER
            cell.font      = LABEL_FONT
            cell.border    = _border()
            cell.fill      = INC_FILL

        # 年度數值
        for ci, y in enumerate(ALL_YEARS, 4):
            rv = roe.get(y)
            rc = ws.cell(row=row, column=ci, value=rv)
            rc.number_format = "0.00"
            rc.alignment     = RIGHT
            rc.border        = _border()
            rc.font          = DATA_FONT
            rc.fill          = ROE_FILL if (rv is None or rv >= 15.0) else WARN_FILL

            iv = inc.get(y)
            ic = ws.cell(row=row+1, column=ci, value=iv)
            ic.number_format = "#,##0.00"
            ic.alignment     = RIGHT
            ic.border        = _border()
            ic.font          = DATA_FONT
            ic.fill          = INC_FILL

        row += 2

    # 底部說明
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=n_cols)
    ws.cell(row=row+1, column=1,
            value="※ 資料來源：永豐金股市資訊網    ※ ROE(A) 底色標黃表示低於 15%"
                  "    ※ 淨利單位：億元（百萬元 ÷ 100）"
            ).font = Font(italic=True, color="757575", name="微軟正黑體", size=9)


# ════════════════════════════════════════════════════════
#  工作表 3：價值低估區
# ════════════════════════════════════════════════════════

def build_value_zone_sheet(ws, all_data: list, prices: dict[str, float | None],
                           price_date: str):
    """
    篩選條件：收盤價 ≤ 合理價值 / 1.2
    潛在獲利比率 = (合理價值/1.2 - 收盤價) / 收盤價
    依潛在獲利比率由高到低排列
    """
    ws.title = "價值低估區"
    ws.freeze_panes = "C3"

    SAFE_FILL  = _clr("E8F5E9")   # 淡綠 — 在低估區內
    RATIO_FILL = _clr("A5D6A7")   # 深綠 — 潛在獲利比率欄

    n_cols = 7

    # 第 1 列：大標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1,
                   value=f"價值低估區（收盤價 ≤ 合理買入價 = 合理價值÷1.2）  報價日期：{price_date}")
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # 第 2 列：欄位標題
    col_headers = ["代號", "名稱", "收盤價", "合理買入價\n(合理價值÷1.2)",
                   "合理價值", "潛在獲利比率\n(合理價值-收盤價)÷收盤價", "近5年\n平均ROE%"]
    col_widths   = [8, 12, 10, 14, 10, 16, 10]
    for i, (h, w) in enumerate(zip(col_headers, col_widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font      = YR_FONT
        cell.fill      = YR_FILL
        cell.alignment = CENTER
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 30

    # 篩選 + 計算
    candidates = []
    for d in all_data:
        sid       = d["stock_id"]
        name      = d["name"]
        value     = d["value"]
        avg_roe   = d["avg_roe_pct"]
        price     = prices.get(sid)

        if value is None or price is None:
            continue

        buy_target = round(value / 1.2, 2)
        if price <= buy_target:
            ratio = round((value - price) / price * 100, 2)
            candidates.append({
                "stock_id":   sid,
                "name":       name,
                "price":      price,
                "buy_target": buy_target,
                "value":      value,
                "ratio":      ratio,
                "avg_roe":    avg_roe,
            })

    # 依潛在獲利比率由高到低排列
    candidates.sort(key=lambda x: x["ratio"], reverse=True)

    # 資料列
    for ri, c in enumerate(candidates, 3):
        row_data = [c["stock_id"], c["name"], c["price"], c["buy_target"],
                    c["value"], c["ratio"], c["avg_roe"]]
        for ci, v in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border    = _border()
            cell.alignment = CENTER if ci <= 2 else RIGHT
            cell.font      = LABEL_FONT if ci <= 2 else DATA_FONT
            if ci <= 2:
                cell.fill = SAFE_FILL
            elif ci == 6:   # 潛在獲利比率
                cell.number_format = "0.00\"%\""
                cell.fill          = RATIO_FILL
            elif ci in (3, 4, 5):
                cell.number_format = "#,##0.00"
                cell.fill          = SAFE_FILL
            else:
                cell.number_format = "0.00"
                cell.fill          = SAFE_FILL
        ws.row_dimensions[ri].height = 18

    # 底部統計
    note_row = len(candidates) + 4
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=n_cols)
    ws.cell(row=note_row, column=1,
            value=(f"共篩出 {len(candidates)} 支 / {len(all_data)} 支    "
                   "※ 潛在獲利比率 = (合理價值 − 收盤價) ÷ 收盤價    "
                   "※ 合理買入價 = 合理價值 ÷ 1.2（留 20% 安全邊際）    "
                   "※ 股價來源：永豐金 zca 頁面（前日收盤）")
            ).font = Font(italic=True, color="757575", name="微軟正黑體", size=9)

    if not candidates:
        ws.cell(row=3, column=1,
                value="目前無股票進入低估區").font = Font(color="757575", name="微軟正黑體", size=11)


# ════════════════════════════════════════════════════════
#  主程式
# ════════════════════════════════════════════════════════

def main():
    with open(VALUE_LIST, "r", encoding="utf-8") as f:
        vl = json.load(f)
    stocks = sorted(vl["stocks"], key=lambda x: x["stock_id"])
    print(f"共 {len(stocks)} 支股票，開始抓取資料…")

    all_data = []
    for i, s in enumerate(stocks, 1):
        sid  = s["stock_id"]
        name = s["name"]
        print(f"  [{i:>2}/{len(stocks)}] {sid} {name}", end="  ", flush=True)

        # 抓取 zcra（ROE / BV）
        zcra = fetch_zcra(sid)
        time.sleep(SLEEP)

        # 抓取 zcqa（歸母淨利）
        inc = fetch_income(sid)
        time.sleep(SLEEP)

        # 抓取 zca（最高/最低本益比 + 前日收盤價，全來自永豐金）
        zca = fetch_zca_data(sid)
        time.sleep(SLEEP)

        # 計算 PE 指標（近 5 年）
        pe_metrics = calc_pe_metrics(zca["high_pe"], zca["low_pe"], PE_YEARS)

        # 近 5 年平均 ROE
        roe_vals = [zcra["roe"].get(y) for y in PE_YEARS
                    if zcra["roe"].get(y) is not None]
        avg_roe = round(sum(roe_vals) / len(roe_vals), 2) if roe_vals else None

        # 最新年度每股淨值
        bv_dict   = zcra["bv"]
        bv_years  = sorted([y for y in bv_dict if bv_dict[y] is not None], reverse=True)
        bv_latest = bv_dict[bv_years[0]] if bv_years else None
        bv_year   = bv_years[0] if bv_years else None

        # 合理價值
        value = calc_value(bv_latest, pe_metrics["max_pe"],
                           pe_metrics["min_pe"], avg_roe)

        n_pe = pe_metrics["n_years"]

        all_data.append({
            "stock_id":  sid,
            "name":      name,
            "roe":       {y: zcra["roe"].get(y) for y in ALL_YEARS},
            "income":    {y: inc.get(y)          for y in ALL_YEARS},
            "bv_latest": bv_latest,
            "bv_year":   bv_year,
            "max_pe":    pe_metrics["max_pe"],
            "min_pe":    pe_metrics["min_pe"],
            "avg_roe_pct": avg_roe,
            "value":     value,
            "n_pe_years": n_pe,
            "price":     zca["price"],    # 前日收盤（來自永豐金 zca 頁面）
        })

        print(f"BV({bv_year})={bv_latest}  高PE={pe_metrics['max_pe']}  "
              f"低PE={pe_metrics['min_pe']}  ROE均={avg_roe}%  合理價={value}")

    # ── 彙整收盤價（已於 zca 抓取時一併取得）──
    import datetime
    price_date = datetime.date.today().strftime("%Y-%m-%d")
    prices = {d["stock_id"]: d["price"] for d in all_data}
    found  = sum(1 for v in prices.values() if v is not None)
    print(f"\n收盤價（永豐金 zca 前日收盤）取得 {found}/{len(all_data)} 支")
    for d in all_data:
        sid = d["stock_id"]
        p   = d["price"]
        val = d["value"]
        buy = round(val / 1.2, 1) if val else None
        zone = "★低估" if (p and buy and p <= buy) else ""
        print(f"  {sid} {d['name']}: 收盤={p}  合理買入={buy}  {zone}")

    # ── 建立 Excel ──
    wb = Workbook()

    # 工作表 1：價值低估區
    ws0 = wb.active
    build_value_zone_sheet(ws0, all_data, prices, price_date)

    # 工作表 2：合理估值
    ws1 = wb.create_sheet()
    build_valuation_sheet(ws1, all_data)

    # 工作表 3：ROE & 歸母淨利
    ws2 = wb.create_sheet()
    build_detail_sheet(ws2, all_data)

    os.makedirs(os.path.dirname(OUTPUT_XLS), exist_ok=True)
    wb.save(OUTPUT_XLS)

    # ── 輸出 value_zone.json（供 notify_value.py 讀取）──
    zone_stocks = []
    for d in all_data:
        val   = d["value"]
        price = prices.get(d["stock_id"])
        if val is None or price is None:
            continue
        buy_target = round(val / 1.2, 2)
        if price <= buy_target:
            zone_stocks.append({
                "stock_id":   d["stock_id"],
                "name":       d["name"],
                "price":      price,
                "buy_target": buy_target,
                "value":      val,
                "ratio":      round((val - price) / price * 100, 2),
                "avg_roe":    d["avg_roe_pct"],
            })
    zone_stocks.sort(key=lambda x: x["ratio"], reverse=True)

    with open(OUTPUT_ZONE, "w", encoding="utf-8") as f:
        json.dump({
            "date":   price_date,
            "total":  len(all_data),
            "stocks": zone_stocks,
        }, f, ensure_ascii=False, indent=2)

    zone_cnt = len(zone_stocks)
    print(f"\n✅ 已儲存 → {OUTPUT_XLS}")
    print(f"   工作表1：價值低估區（{zone_cnt} 支）")
    print(f"   工作表2：合理估值（{len(stocks)} 支）")
    print(f"   工作表3：ROE & 歸母淨利（{len(stocks)} 支 × 2 指標 × 8 年）")
    print(f"✅ 已儲存 → {OUTPUT_ZONE}（低估區 {zone_cnt} 支）")


if __name__ == "__main__":
    main()
